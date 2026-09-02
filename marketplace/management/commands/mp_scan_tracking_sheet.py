"""Scan a Flipkart scanning sheet into Outward from the server side.

Same job as the Outward page's "upload a scanning sheet" control, but run on the
box instead of through the browser — useful when the sheet is long enough that the
operator does not want to sit on a spinner, or when the page cannot reach the file.

It calls the SAME service the barcode gun and the bulk endpoint call
(``scan_dispatch_by_tracking``), one ID per transaction, so packing / cancelled /
unmapped-SKU rules still decide what ships and one bad row never rolls back the
good ones. Nothing is written without ``--apply``.

The dry run is the diagnosis. It puts every ID through that same service and then
rolls the transaction back, so it reports the REAL verdict — scanned, already
scanned, or refused with the reason — while leaving the database untouched. That
is what answers "Flipkart says 443 parcels, the board shows 417": the shortfall is
listed with a reason against each ID, and ``--out`` writes the lot to a CSV. It
deliberately re-uses the live rules rather than re-implementing them, because a
second copy of the rules would drift and quietly lie about what will happen.

    # diagnose — writes nothing
    python manage.py mp_scan_tracking_sheet "Flipkart Scanning 01 09 2026.xlsx" \\
        --company JIVO_MART --out shortfall.csv

    # then record them
    python manage.py mp_scan_tracking_sheet sheet.xlsx --company JIVO_MART --apply
    python manage.py mp_scan_tracking_sheet sheet.csv --company JIVO_MART --apply --user ops@jivo.in
"""
import csv
from collections import Counter

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from accounts.models import User
from company.models import Company
from marketplace.models import MarketplaceChannel
from marketplace.services.errors import MarketplaceError
from marketplace.services.scan_service import scan_dispatch_by_tracking

# A first-row cell that is the column title rather than a shipment.
HEADER_CELL = {"tracking id", "trackingid", "tracking ids", "tracking", "tracking no",
               "tracking number", "tracking_id"}


def read_tracking_ids(path):
    """First column of the first sheet/CSV, de-duplicated, header row dropped."""
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - openpyxl ships with the app
            raise CommandError("openpyxl is required to read .xlsx sheets") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cells = [row[0] if row else None for row in ws.iter_rows(values_only=True)]
    else:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            cells = [row[0] if row else "" for row in csv.reader(fh)]

    ids, seen = [], set()
    for index, value in enumerate(cells):
        code = str(value or "").strip()
        if not code:
            continue
        if index == 0 and code.lower() in HEADER_CELL:
            continue
        key = code.upper()
        if key in seen:
            continue
        seen.add(key)
        ids.append(code)
    return ids


class _DryRun(Exception):
    """Raised once a dry-run ID's verdict is known, to unwind its transaction.

    Carries the verdict out of the block it destroys, so a dry run can report the
    real outcome of every ID without leaving a single row behind.
    """

    def __init__(self, duplicate):
        super().__init__("dry run")
        self.duplicate = duplicate


class Command(BaseCommand):
    help = "Scan every Tracking ID in a scanning sheet into Outward (dry run unless --apply)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx / .csv scanning sheet")
        parser.add_argument("--company", required=True, help="Company code, e.g. JIVO_MART")
        parser.add_argument("--channel", default=MarketplaceChannel.FLIPKART)
        parser.add_argument("--user", default="", help="Email to record as the scanner")
        parser.add_argument("--apply", action="store_true", help="Actually record the scans")
        parser.add_argument(
            "--out", default="",
            help="Write every ID that did not scan to this CSV (tracking_id, reason, message)",
        )

    def handle(self, *args, **opts):
        company = Company.objects.filter(code=opts["company"]).first()
        if company is None:
            raise CommandError(
                f"No company with code {opts['company']!r}. Known: "
                + ", ".join(Company.objects.values_list("code", flat=True))
            )
        user = None
        if opts["user"]:
            user = User.objects.filter(email=opts["user"]).first()
            if user is None:
                raise CommandError(f"No user with email {opts['user']!r}")

        ids = read_tracking_ids(opts["path"])
        db = connection.settings_dict
        apply = opts["apply"]
        self.stdout.write(f"DB      : {db['NAME']}@{db['HOST']}")
        self.stdout.write(f"Company : {company.code} ({company.name})")
        self.stdout.write(f"Channel : {opts['channel']}")
        self.stdout.write(f"Scan as : {getattr(user, 'email', None) or '(unattributed)'}")
        self.stdout.write(f"Sheet   : {opts['path']} -> {len(ids)} tracking IDs")
        if not ids:
            raise CommandError("No tracking IDs found in the first column of that sheet.")
        if not apply:
            self.stdout.write(self.style.WARNING(
                "Dry run - every ID is put through the real scan and then rolled back, "
                "so the verdicts below are exact and nothing is written."
            ))

        scanned = already = failed = 0
        reasons = Counter()
        refusals = []
        for n, code in enumerate(ids, 1):
            try:
                with transaction.atomic():
                    _dispatch, _created, is_duplicate = scan_dispatch_by_tracking(
                        company, opts["channel"], barcode=code, user=user,
                    )
                    if not apply:
                        # Verdict reached; abandon the write. Raising unwinds this
                        # atomic block (a savepoint inside the service's own
                        # transaction), so the dry run exercises the REAL rules
                        # rather than a second copy of them that could drift.
                        raise _DryRun(is_duplicate)
            except _DryRun as dry:
                if dry.duplicate:
                    already += 1
                    reasons["ALREADY_SCANNED"] += 1
                else:
                    scanned += 1
            except MarketplaceError as exc:
                failed += 1
                reason = getattr(exc, "code", "") or "REFUSED"
                reasons[reason] += 1
                refusals.append((code, reason, str(exc)))
            except Exception as exc:  # noqa: BLE001 - one bad row must not stop the sheet
                failed += 1
                reasons[type(exc).__name__] += 1
                refusals.append((code, type(exc).__name__, str(exc)))
            else:
                if is_duplicate:
                    already += 1
                    reasons["ALREADY_SCANNED"] += 1
                else:
                    scanned += 1
            if n % 25 == 0 or n == len(ids):
                self.stdout.write(
                    f"  {n}/{len(ids)}  scanned={scanned} already={already} failed={failed}",
                    ending="\n",
                )
                self.stdout.flush()

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"scanned {scanned} | already scanned {already} | not scanned {failed} "
            f"| total {len(ids)}"
        ))
        for reason, count in reasons.most_common():
            self.stdout.write(f"  {reason}: {count}")
        if refusals:
            self.stdout.write("")
            self.stdout.write("Not scanned (first 30):")
            for code, reason, message in refusals[:30]:
                self.stdout.write(f"  {code}  [{reason}] {message}")
            if len(refusals) > 30 and not opts["out"]:
                self.stdout.write(
                    f"  ... and {len(refusals) - 30} more — re-run with "
                    f"--out refusals.csv to get all of them."
                )

        if opts["out"]:
            # The whole shortfall as a file, so the sheet Flipkart says it handed over
            # can be reconciled against what the app holds without reading a terminal.
            with open(opts["out"], "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(["Tracking ID", "Reason", "Message"])
                writer.writerows(refusals)
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS(
                f"Wrote {len(refusals)} unscanned tracking ID(s) to {opts['out']}"
            ))
