"""Amazon order-sheet parser — deliberately SEPARATE from the Flipkart parser.

Amazon's tax/shipment report (MTR) has an entirely different layout from Flipkart's
order CSV (84 columns, ASIN instead of FSN, Shipment Id instead of a Tracking ID,
Transaction Type instead of Order State, and it ships as .xlsx). Keeping this in its
own module means a change to Amazon's column mapping can never affect Flipkart, and
vice-versa. It emits the SAME canonical row dicts that ``order_import_service``
consumes, so the rest of the flow (resolve → issue → pack → scan → confirm → DN) is
shared and channel-scoped.

Accepts both ``.xlsx`` and ``.csv``. Excel parsing uses ``openpyxl`` when present and
otherwise a small standard-library reader (an ``.xlsx`` is a zip of XML), so it works
on any server without an extra dependency.
"""
import csv
import io
import re
import zipfile
from datetime import date, datetime, timedelta
from xml.etree import ElementTree as ET

from .errors import MarketplaceError

# Canonical field → Amazon column label. Matching is case/space-insensitive.
# The canonical keys mirror Flipkart's so `ingest()` reads the same shape.
AMAZON_COLUMNS = {
    "order_id": "Order Id",
    "order_item_id": "Shipment Item Id",
    "shipment_id": "Shipment Id",
    "ordered_on": "Order Date",
    "order_state": "Transaction Type",
    "order_type": "Fulfillment Channel",
    "fsn": "Asin",                 # Amazon's product key (used like Flipkart's FSN)
    "sku": "Sku",
    "product": "Item Description",
    "quantity": "Quantity",
    "hsn": "Hsn/sac",
    "unit_price": "Principal Amount",
    "invoice_amount": "Invoice Amount",
    "cgst": "Cgst Tax",
    "igst": "Igst Tax",
    "sgst": "Sgst Tax",
    "buyer": "Ship To City",       # MTR carries no buyer name; ship-to city is the label
    "ship_to": "Ship To City",
    "addr1": "",
    "addr2": "",
    "city": "Ship To City",
    "state": "Ship To State",
    "pin": "Ship To Postal Code",
    "dispatch_by": "Shipment Date",
    "tracking": "Shipment Id",      # the scannable per-shipment id (Amazon's "Tracking ID")
}

# Transaction Types that mean the order was cancelled/returned (not to be fulfilled).
_CANCEL_TYPES = {"cancel", "refund", "return"}


def _norm(s):
    return (s or "").strip().lower().replace("_", " ").replace("  ", " ")


def _cell(value):
    """Stringify one cell. Dates are rendered in Flipkart's date style so the shared
    ``_parse_dt`` reads them without touching its format list."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%b %d, %Y %H:%M:%S") if isinstance(value, datetime) else value.strftime("%b %d, %Y")
    return str(value).strip()


def _is_xlsx(content, filename, content_type):
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return True
    if name.endswith(".csv"):
        return False
    if "sheet" in (content_type or "").lower() or "excel" in (content_type or "").lower():
        return True
    # xlsx files are ZIP archives → magic bytes "PK".
    return bool(content) and content[:2] == b"PK"


def _local(tag):
    return tag.rsplit("}", 1)[-1]  # strip XML namespace


def _col_index(ref):
    """'AB12' → zero-based column index (A→0, B→1, … AA→26)."""
    letters = re.match(r"[A-Za-z]+", ref or "")
    if not letters:
        return 0
    n = 0
    for ch in letters.group(0).upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _maybe_excel_date(value):
    """A bare Excel serial number in a date column → a Flipkart-style date string.

    Only used for the date fields, and only when a value is purely numeric (the
    stdlib reader can't apply cell formatting the way openpyxl does)."""
    if not value:
        return value
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return value  # already a formatted date string
    if serial <= 0:
        return value
    try:
        return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime("%b %d, %Y %H:%M:%S")
    except (OverflowError, ValueError):
        return value


def _rows_from_xlsx_stdlib(content):
    """Read the first worksheet of an .xlsx using only the standard library.

    An .xlsx is a zip of XML: ``sharedStrings.xml`` (the text table) + a worksheet
    with cells that either carry a shared-string index (``t="s"``), an inline string,
    or a raw number. Dates arrive as Excel serial numbers (converted later for the
    date fields)."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        raise MarketplaceError("The Excel file could not be read.", code="BAD_SHEET", status_code=400)

    shared = []
    if "xl/sharedStrings.xml" in zf.namelist():
        for si in ET.fromstring(zf.read("xl/sharedStrings.xml")):
            if _local(si.tag) == "si":
                shared.append("".join(t.text or "" for t in si.iter() if _local(t.tag) == "t"))

    sheet = next((n for n in zf.namelist()
                  if n.startswith("xl/worksheets/") and n.endswith(".xml")), None)
    if sheet is None:
        return []

    rows = []
    for el in ET.fromstring(zf.read(sheet)).iter():
        if _local(el.tag) != "row":
            continue
        cells, max_col = {}, -1
        for c in el:
            if _local(c.tag) != "c":
                continue
            ci = _col_index(c.get("r", ""))
            max_col = max(max_col, ci)
            value, ctype = None, c.get("t")
            for child in c:
                lt = _local(child.tag)
                if lt == "v":
                    value = child.text
                elif lt == "is":  # inline string
                    value = "".join(x.text or "" for x in child.iter() if _local(x.tag) == "t")
            if ctype == "s" and value is not None:
                try:
                    value = shared[int(value)]
                except (ValueError, IndexError):
                    value = ""
            cells[ci] = value if value is not None else ""
        rows.append([str(cells.get(i, "")).strip() for i in range(max_col + 1)])
    return rows


def _rows_from_xlsx(content):
    """Prefer openpyxl (handles cell formatting/dates natively); fall back to the
    standard-library reader when openpyxl isn't installed on the server."""
    try:
        import openpyxl
    except Exception:  # openpyxl not available — use the stdlib reader
        return _rows_from_xlsx_stdlib(content)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]


def _raw_rows(*, text, content, filename, content_type):
    """Return a list of raw rows (list of stringified cells); row 0 is the header."""
    if _is_xlsx(content, filename, content_type):
        if content is None:
            raise MarketplaceError("Excel upload is empty.", code="BAD_SHEET", status_code=400)
        return _rows_from_xlsx(content)
    # CSV
    if text is None and content is not None:
        text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text or ""))
    return [[_cell(c) for c in row] for row in reader]


def parse_amazon_rows(*, text=None, content=None, filename="", content_type=""):
    """Parse an Amazon sheet into canonical row dicts (same keys as the Flipkart
    parser). Raises ``BAD_SHEET`` if the required columns are absent."""
    raw = _raw_rows(text=text, content=content, filename=filename, content_type=content_type)
    if not raw:
        raise MarketplaceError("The sheet is empty.", code="BAD_SHEET", status_code=400)

    header = raw[0]
    index = {_norm(h): i for i, h in enumerate(header)}
    col = {key: index.get(_norm(label)) for key, label in AMAZON_COLUMNS.items() if label}
    if col.get("order_id") is None or col.get("sku") is None or col.get("quantity") is None:
        raise MarketplaceError(
            "Amazon sheet is missing required columns (Order Id, Sku, Quantity).",
            code="BAD_SHEET", status_code=400,
        )

    rows = []
    for raw_row in raw[1:]:
        if not any((c or "").strip() for c in raw_row):
            continue

        def get(key):
            i = col.get(key)
            return raw_row[i].strip() if i is not None and i < len(raw_row) and raw_row[i] else ""

        row = {key: get(key) for key in AMAZON_COLUMNS}
        # Dates: the stdlib reader yields Excel serial numbers — convert them (a
        # no-op when openpyxl already produced a formatted date string).
        row["ordered_on"] = _maybe_excel_date(row["ordered_on"])
        row["dispatch_by"] = _maybe_excel_date(row["dispatch_by"])
        # Normalise the state so the shared cancellation check (``"cancel" in state``)
        # works without changing Flipkart's logic.
        if _norm(row["order_state"]) in _CANCEL_TYPES:
            row["order_state"] = "Cancelled"
        rows.append(row)
    return rows
