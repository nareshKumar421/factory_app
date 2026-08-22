"""
sap_reports/exports.py

Turning a report result into a file.

The people who run these reports in SAP export them to Excel, so the app has to
do the same or it is a step backwards. Both writers stream row by row: an export
is allowed ten times the on-screen row ceiling, which is enough to matter.
"""

import csv
import io
import re
from typing import Dict, List

from django.http import HttpResponse
from django.utils import timezone

CSV_CONTENT_TYPE = "text/csv"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Excel refuses these in a sheet name, and truncates past 31 characters.
INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]")


def build_filename(title: str, extension: str) -> str:
    """``Pending Dispatch`` -> ``pending-dispatch-20260822-1431.xlsx``."""
    stem = re.sub(r"[^A-Za-z0-9]+", "-", title).strip("-").lower() or "sap-report"
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"{stem}-{stamp}.{extension}"


def csv_response(title: str, columns: List[Dict], rows: List[List]) -> HttpResponse:
    """The result as a CSV download."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([column["label"] for column in columns])
    for row in rows:
        writer.writerow(["" if value is None else value for value in row])

    # utf-8-sig so Excel opens non-ASCII headings correctly on Windows.
    response = HttpResponse(
        buffer.getvalue().encode("utf-8-sig"),
        content_type=CSV_CONTENT_TYPE,
    )
    _attach(response, build_filename(title, "csv"))
    return response


def xlsx_response(
    title: str,
    columns: List[Dict],
    rows: List[List],
    meta: Dict = None,
) -> HttpResponse:
    """
    The result as an .xlsx download, with the filters recorded on a second sheet.

    A report exported without its filters is a number without a question, and
    these get mailed around; the "Filters" sheet is what makes the file readable
    a week later.
    """
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet(_sheet_name(title))

    header_font = Font(bold=True)
    header = []
    for column in columns:
        cell = openpyxl.cell.WriteOnlyCell(sheet, value=column["label"])
        cell.font = header_font
        header.append(cell)
    sheet.append(header)

    for row in rows:
        sheet.append([_excel_safe(value) for value in row])

    if meta:
        _append_filter_sheet(workbook, meta)

    buffer = io.BytesIO()
    workbook.save(buffer)

    response = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    _attach(response, build_filename(title, "xlsx"))
    return response


def _append_filter_sheet(workbook, meta: Dict) -> None:
    from openpyxl.styles import Font

    sheet = workbook.create_sheet("Filters")
    bold = Font(bold=True)

    def labelled(label, value):
        import openpyxl

        cell = openpyxl.cell.WriteOnlyCell(sheet, value=label)
        cell.font = bold
        sheet.append([cell, value])

    labelled("Report", meta.get("title", ""))
    labelled("Company", meta.get("company", ""))
    labelled("Run at", meta.get("executed_at", ""))
    labelled("Rows", meta.get("row_count", 0))
    if meta.get("was_truncated"):
        labelled("Note", f"Cut off at the {meta.get('row_limit')} row limit — narrow the filters.")

    sheet.append([])
    for parameter in meta.get("parameters", []) or []:
        labelled(parameter.get("label", ""), parameter.get("value", ""))


def _sheet_name(title: str) -> str:
    cleaned = INVALID_SHEET_CHARS_RE.sub(" ", title).strip() or "Report"
    return cleaned[:31]


def _excel_safe(value):
    """
    Keeps openpyxl from choking, and keeps Excel from running a cell as a formula.

    SAP item names and remarks occasionally start with ``=`` or ``-``; prefixing
    an apostrophe is the standard way to make such a cell inert.
    """
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    return f"'{text}" if text[:1] in ("=", "+", "@") else text


def _attach(response: HttpResponse, filename: str) -> None:
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    # So a browser fetch() can read the name off the response.
    response["Access-Control-Expose-Headers"] = "Content-Disposition"
